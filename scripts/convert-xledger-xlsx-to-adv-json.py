import json
import re
import sys
from pathlib import Path

import openpyxl


FIELD_TYPES = {
    "Value": "Integer (bigint)",
    "i_level": "Integer",
    "Kode": "Text (nvarchar)",
    "s_combined": "Text (nvarchar)",
    "Beskrivelse": "Text (nvarchar)",
    "Navn": "Text (nvarchar)",
    "Belop": "Money",
    "Beløp": "Money",
}


MONTHS = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "Mai": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Okt": 10,
    "Nov": 11,
    "Des": 12,
}


def cell_value(value):
    if value is None:
        return ""
    return value


def period_for_title(title):
    text = str(title or "")
    match = re.search(r"(IB|Jan|Feb|Mar|Apr|Mai|Jun|Jul|Aug|Sep|Okt|Nov|Des)/(\d{4})", text)
    if not match:
        return None
    label, year = match.groups()
    period = 0 if label == "IB" else MONTHS[label]
    return {
        "fiscal_year": int(year),
        "period": period,
        "description": text.replace("..", ""),
    }


def build_field_groups(group_titles, field_titles):
    groups = []
    current_group = None
    current_fields = []
    current_period = None

    def flush():
        nonlocal current_group, current_fields, current_period
        if not current_group:
            return
        group = {
            "title": current_group,
            "fields": current_fields,
        }
        if current_period:
            group["periods"] = [current_period]
        groups.append(group)

    for group_title, field_title in zip(group_titles, field_titles):
        if group_title:
            flush()
            current_group = str(group_title)
            current_fields = []
            current_period = period_for_title(current_group)
        if current_group:
            title = str(field_title or "Value")
            current_fields.append({
                "title": title,
                "type": FIELD_TYPES.get(title, "Text (nvarchar)")
            })

    flush()
    return groups


def convert(input_path, output_path):
    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook["Data"] if "Data" in workbook.sheetnames else workbook.active
    rows = worksheet.iter_rows(values_only=True)
    group_titles = [cell_value(value) for value in next(rows)]
    field_titles = [cell_value(value) for value in next(rows)]
    field_groups = build_field_groups(group_titles, field_titles)

    group_widths = [len(group["fields"]) for group in field_groups]
    output_rows = []
    for source_row in rows:
        cursor = 0
        target_row = []
        for width in group_widths:
            values = [cell_value(value) for value in source_row[cursor:cursor + width]]
            target_row.append(values)
            cursor += width
        output_rows.append(target_row)

    result = {
        "meta": {
            "rows_returned": len(output_rows),
            "converted_from": str(input_path),
        },
        "field_groups": field_groups,
        "rows": output_rows,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Usage: convert-xledger-xlsx-to-adv-json.py input.xlsx output.adv.json")
    convert(Path(sys.argv[1]), Path(sys.argv[2]))
